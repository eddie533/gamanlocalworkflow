#!/usr/bin/env python3
"""
Stage 2 - Enrichment Pipeline

Reads an input Excel (Companies + People sheets), enriches companies with
Vertical/Sub-Vertical/Informal Name via GPT-5, guesses contact gender,
and outputs a transformed Excel matching the target format.
"""

import asyncio
import sys
from datetime import date
from pathlib import Path

import openpyxl

from enrich import CompanyEnricher
from gender import GenderGuesser
from role_classifier import RoleClassifier
from city_extractor import CityExtractor

MAX_WORKERS = 50
INPUT_DIR = Path(__file__).parent / "input"
OUTPUT_DIR = Path(__file__).parent / "output"


# ── helpers ──────────────────────────────────────────────────────────────

def find_input_file() -> Path:
    """Return the first .xlsx in INPUT_DIR, or accept a CLI argument."""
    if len(sys.argv) > 1:
        return Path(sys.argv[1])
    xlsx_files = sorted(INPUT_DIR.glob("*.xlsx"))
    if not xlsx_files:
        print(f"No .xlsx files found in {INPUT_DIR}")
        sys.exit(1)
    return xlsx_files[0]


def read_sheet(wb: openpyxl.Workbook, name: str) -> list[dict]:
    """Read a worksheet into a list of dicts keyed by header row."""
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [h for h in rows[0] if h is not None]
    return [
        {headers[i]: cell for i, cell in enumerate(row[:len(headers)])}
        for row in rows[1:]
        if any(cell is not None for cell in row)
    ]


# ── async enrichment ────────────────────────────────────────────────────

async def enrich_one(company: dict, semaphore: asyncio.Semaphore):
    async with semaphore:
        enricher = CompanyEnricher()
        result = await asyncio.to_thread(enricher.enrich_company, company)
        return result


async def enrich_all(companies: list[dict]) -> list[dict]:
    semaphore = asyncio.Semaphore(MAX_WORKERS)
    tasks = [enrich_one(c, semaphore) for c in companies]
    return await asyncio.gather(*tasks, return_exceptions=True)


# ── async gender guessing ───────────────────────────────────────────────

async def guess_one(first_name: str, semaphore: asyncio.Semaphore):
    async with semaphore:
        guesser = GenderGuesser()
        result = await asyncio.to_thread(guesser.guess, first_name)
        return result


async def guess_all(first_names: list[str]) -> list[str]:
    semaphore = asyncio.Semaphore(MAX_WORKERS)
    tasks = [guess_one(name, semaphore) for name in first_names]
    return await asyncio.gather(*tasks, return_exceptions=True)


# ── async role classification ────────────────────────────────────────────

async def classify_one(role: str, semaphore: asyncio.Semaphore):
    async with semaphore:
        classifier = RoleClassifier()
        result = await asyncio.to_thread(classifier.classify, role)
        return result


async def classify_all(roles: list[str]) -> list[str]:
    semaphore = asyncio.Semaphore(MAX_WORKERS)
    tasks = [classify_one(role, semaphore) for role in roles]
    return await asyncio.gather(*tasks, return_exceptions=True)


# ── async city extraction ────────────────────────────────────────────────

async def extract_city_one(address: str, semaphore: asyncio.Semaphore):
    async with semaphore:
        extractor = CityExtractor()
        result = await asyncio.to_thread(extractor.extract, address)
        return result


async def extract_cities(addresses: list[str]) -> list[str]:
    semaphore = asyncio.Semaphore(MAX_WORKERS)
    tasks = [extract_city_one(addr, semaphore) for addr in addresses]
    return await asyncio.gather(*tasks, return_exceptions=True)


# ── output builder ───────────────────────────────────────────────────────

COMPANY_COLUMNS = [
    "Company Name", "Country", "Company Website", "Vertical", "Sub-Vertical",
    "Date Added", "Added By", "Legal name", "Source", "Status",
    "Founding Year", "City", "LinkedIn Page", "Currency", "FTEs",
]

CONTACT_COLUMNS = [
    "Company", "First Name", "Last Name", "Company Website",
    "LinkedIn", "Gender", "Role", "Age", "Email", "Status",
]


def build_output(
    enrichment_results: list[dict],
    raw_companies: list[dict],
    city_results: list[str],
    people: list[dict],
    gender_results: list[str],
    role_results: list[str],
    output_path: Path,
):
    """Write the two-sheet output Excel."""
    # Build lookups: original company name → enrichment result / raw input
    enrich_map: dict[str, dict] = {}
    raw_map: dict[str, dict] = {}
    for r in enrichment_results:
        if isinstance(r, Exception):
            continue
        enrich_map[r["company_name"]] = r
    for c in raw_companies:
        raw_map[c.get("Company Name", "")] = c

    today = date.today().isoformat()

    wb = openpyxl.Workbook()

    # ── Companies sheet ──
    ws_comp = wb.active
    ws_comp.title = "Companies"
    ws_comp.append(COMPANY_COLUMNS)

    for r, city in zip(enrichment_results, city_results):
        if isinstance(r, Exception):
            continue
        raw = raw_map.get(r["company_name"], {})
        cd = r["company_data"]
        ws_comp.append([
            r.get("informal_name", ""),                              # Company Name
            raw.get("Country", ""),                                  # Country
            cd.get("Website URL", cd.get("Website", "")),            # Company Website
            r.get("vertical", ""),                                   # Vertical
            r.get("subvertical", ""),                                # Sub-Vertical
            today,                                                   # Date Added
            "Eddie Childs",                                          # Added By
            r.get("company_name", ""),                               # Legal name
            "Proprietary",                                           # Source
            "Reach Out",                                             # Status
            raw.get("Founding Year", ""),                             # Founding Year
            city if not isinstance(city, Exception) else "",          # City
            raw.get("LinkedIn URL", ""),                              # LinkedIn Page
            "",                                                      # Currency
            raw.get("FTE Count", ""),                                 # FTEs
        ])

    # ── Contacts sheet ──
    ws_cont = wb.create_sheet("Contacts")
    ws_cont.append(CONTACT_COLUMNS)

    for person, gender, role in zip(people, gender_results, role_results):
        if isinstance(gender, Exception):
            gender = "Unknown"
        if isinstance(role, Exception):
            role = "Management NOT Founder/Shareholder"

        company_name = person.get("Company Name", "")
        enriched = enrich_map.get(company_name, {})
        informal = enriched.get("informal_name", company_name)

        ws_cont.append([
            informal,                               # Company
            person.get("First Name", ""),            # First Name
            person.get("Last Name", ""),             # Last Name
            person.get("Website URL", ""),           # Company Website
            person.get("LinkedIn", ""),              # LinkedIn
            gender,                                  # Gender
            role,                                    # Role
            "",                                      # Age (placeholder)
            person.get("Email", ""),                 # Email
            "",                                      # Status
        ])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ── main ─────────────────────────────────────────────────────────────────

def main():
    input_file = find_input_file()

    wb = openpyxl.load_workbook(input_file, read_only=True)
    companies = read_sheet(wb, "Companies")
    people = read_sheet(wb, "People")
    wb.close()

    print(f"Stage 2 — {len(companies)} companies, {len(people)} contacts\n")

    # Prepare company dicts for enrich.py
    enrich_inputs = []
    for c in companies:
        enrich_inputs.append({
            "Company Name": c.get("Company Name", ""),
            "Website": c.get("Website URL", ""),
            "Description": c.get("Description", ""),
            "Country": c.get("Country", ""),
        })

    # Enrich companies
    print("Enriching companies...")
    enrichment_results = asyncio.run(enrich_all(enrich_inputs))
    for r in enrichment_results:
        if isinstance(r, Exception):
            print(f"  error: {r}")
        else:
            print(f"  {r['company_name']} -> {r.get('informal_name')} [{r.get('vertical')}]")

    # Extract cities from addresses
    print("\nExtracting cities...")
    addresses = [c.get("Address", "") for c in companies]
    city_results = asyncio.run(extract_cities(addresses))
    for c, city in zip(companies, city_results):
        ct = city if not isinstance(city, Exception) else ""
        print(f"  {c.get('Company Name', '')} -> {ct}")

    # Guess genders
    print("\nGuessing genders...")
    first_names = [p.get("First Name", "") for p in people]
    gender_results = asyncio.run(guess_all(first_names))
    for person, gender in zip(people, gender_results):
        g = gender if not isinstance(gender, Exception) else "Unknown"
        print(f"  {person.get('First Name', '')} -> {g}")

    # Classify roles
    print("\nClassifying roles...")
    raw_roles = [p.get("Roles", "") for p in people]
    role_results = asyncio.run(classify_all(raw_roles))
    for person, role in zip(people, role_results):
        r = role if not isinstance(role, Exception) else "Management NOT Founder/Shareholder"
        print(f"  {person.get('First Name', '')} ({person.get('Roles', '')}) -> {r}")

    # Build output
    output_path = OUTPUT_DIR / f"enriched_{input_file.stem}.xlsx"
    build_output(enrichment_results, companies, city_results, people, gender_results, role_results, output_path)

    ok = sum(1 for r in enrichment_results if not isinstance(r, Exception))
    errs = len(enrichment_results) - ok
    print(f"\nDone — {ok} companies enriched, {len(people)} contacts processed")
    if errs:
        print(f"  {errs} enrichment errors")
    print(f"Output: {output_path}")


if __name__ == "__main__":
    main()
